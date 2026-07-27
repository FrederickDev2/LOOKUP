"""Excel (.xlsx) import into SQLite.

Uses openpyxl in read-only / streaming mode so that large files (150-200MB,
60,000+ rows) can be imported without loading the whole workbook into memory.

Imports ADD & MERGE into the existing dataset: rows are upserted on the
normalized NIA number, so a NIA already in the database is UPDATED with the
newer values and new NIAs are inserted. The whole import runs inside a single
transaction, so a failed import leaves the previous data completely intact.
"""
from __future__ import annotations

import datetime as _dt
import json
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook

from .columns import NIA_HEADER, header_key
from .database import get_conn, set_meta, utcnow_iso
from .nia import normalize_nia

BATCH_SIZE = 2000

# Insert new members; if the normalized NIA already exists, update it in place.
UPSERT_SQL = (
    "INSERT INTO members (nia_normalized, nia_original, data_json) "
    "VALUES (?, ?, ?) "
    "ON CONFLICT(nia_normalized) DO UPDATE SET "
    "nia_original=excluded.nia_original, data_json=excluded.data_json"
)


@dataclass
class ImportResult:
    row_count: int
    skipped_count: int
    columns: List[str]


def _cell_to_str(value: object) -> str:
    """Render a cell value as a clean string for storage."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (_dt.datetime, _dt.date)):
        # Store dates as ISO date; drop a midnight time component.
        if isinstance(value, _dt.datetime) and value.time() == _dt.time(0, 0):
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, float):
        # Excel stores integers as floats; render whole numbers without ".0".
        if value.is_integer():
            return str(int(value))
        return repr(value)
    return str(value).strip()


def import_excel(path: str, imported_by: str) -> ImportResult:
    """Parse the workbook at `path` and merge (upsert) it into the members table.

    Raises ValueError on a structural problem (missing NIA column, empty sheet).
    """
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)

        try:
            raw_header = next(rows)
        except StopIteration:
            raise ValueError("The uploaded file appears to be empty.")

        headers = [header_key(h) if h is not None else f"COLUMN_{i+1}"
                   for i, h in enumerate(raw_header)]

        # Locate the NIA column.
        try:
            nia_idx = headers.index(NIA_HEADER)
        except ValueError:
            raise ValueError(
                f"Could not find a '{NIA_HEADER}' column. Found columns: "
                + ", ".join(h for h in headers if h) + "."
            )

        row_count = 0
        skipped = 0

        with get_conn() as conn:
            try:
                conn.execute("BEGIN")

                batch: List[tuple] = []
                for raw_row in rows:
                    # Skip entirely-empty rows.
                    if raw_row is None or all(c is None or str(c).strip() == "" for c in raw_row):
                        continue

                    values = list(raw_row)
                    # Pad/trim to header width.
                    if len(values) < len(headers):
                        values += [None] * (len(headers) - len(values))

                    record = {}
                    for i, col in enumerate(headers):
                        if not col:
                            continue
                        record[col] = _cell_to_str(values[i])

                    nia_original = record.get(NIA_HEADER, "")
                    nia_norm = normalize_nia(nia_original)
                    # Rows without a NIA are stored with a NULL key so they are
                    # never merged together and never match a lookup.
                    if not nia_norm:
                        skipped += 1
                        nia_key = None
                    else:
                        nia_key = nia_norm
                    batch.append((nia_key, nia_original, json.dumps(record, ensure_ascii=False)))
                    row_count += 1

                    if len(batch) >= BATCH_SIZE:
                        conn.executemany(UPSERT_SQL, batch)
                        batch.clear()

                if batch:
                    conn.executemany(UPSERT_SQL, batch)
                conn.commit()
            except Exception:
                conn.rollback()
                raise
    finally:
        wb.close()

    if row_count == 0:
        raise ValueError("No data rows were found in the file.")

    set_meta("last_import_at", utcnow_iso())
    set_meta("last_import_rows", str(row_count))
    return ImportResult(row_count=row_count, skipped_count=skipped, columns=headers)


def clear_all_members() -> int:
    """Delete every member record. Returns how many were removed.

    User accounts and audit logs are untouched. This is irreversible (restore
    from a database backup to recover).
    """
    with get_conn() as conn:
        before = conn.execute("SELECT COUNT(*) AS c FROM members").fetchone()["c"]
        conn.execute("DELETE FROM members")
        conn.commit()
    set_meta("last_import_at", "")
    set_meta("last_import_rows", "0")
    return before


def member_count() -> int:
    with get_conn() as conn:
        return conn.execute("SELECT COUNT(*) AS c FROM members").fetchone()["c"]


def last_import_info() -> dict:
    from .database import get_meta
    return {
        "last_import_at": get_meta("last_import_at"),
        "last_import_rows": get_meta("last_import_rows"),
        "member_count": member_count(),
    }
